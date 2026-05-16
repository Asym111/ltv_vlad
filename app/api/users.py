from __future__ import annotations
import io
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import get_db
from app.models.user import User
from app.models.transaction import Transaction
from app.models.bonus_grant import BonusGrant
from app.schemas.user import UserCreate, UserOut, UserUpdate

router = APIRouter(prefix="/users", tags=["users"])


def normalize_phone(raw: str) -> str:
    s = (raw or "").strip()
    s = s.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if s.startswith("+"):
        s = s[1:]
    digits = "".join(ch for ch in s if ch.isdigit())
    if digits.startswith("8") and len(digits) == 11:
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) > 11:
        digits = digits[-11:]
    return digits


def must_tenant_id(request: Request) -> int:
    u = getattr(request.state, "user", None) or {}
    tid = u.get("tenant_id")
    if not tid:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return int(tid)


def require_role(request: Request, *allowed: str):
    u = getattr(request.state, "user", None) or {}
    role = u.get("role", "")
    if role not in allowed:
        raise HTTPException(status_code=403, detail=f"Access denied. Required roles: {allowed}")
    return u


@router.get("/", response_model=list[UserOut])
def list_users(request: Request, db: Session = Depends(get_db)) -> list[UserOut]:
    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin", "manager")
    users = db.query(User).filter(User.tenant_id == tenant_id).order_by(User.id.desc()).all()
    return [UserOut.model_validate(u) for u in users]


@router.post("", response_model=UserOut)
def create_user(payload: UserCreate, request: Request, db: Session = Depends(get_db)) -> UserOut:
    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin", "manager")
    phone = normalize_phone(payload.phone)
    if not phone:
        raise HTTPException(status_code=400, detail="Invalid phone")
    exists = db.query(User).filter(User.tenant_id == tenant_id, User.phone == phone).first()
    if exists:
        raise HTTPException(status_code=400, detail="Phone already exists")
    user = User(
        tenant_id=tenant_id,
        phone=phone,
        full_name=payload.full_name,
        birth_date=payload.birth_date,
        tier=payload.tier or "Bronze",
        bonus_balance=int(payload.bonus_balance or 0),
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return UserOut.model_validate(user)


@router.patch("/{user_id}", response_model=UserOut)
def update_user(user_id: int, payload: UserUpdate, request: Request, db: Session = Depends(get_db)) -> UserOut:
    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin", "manager")
    user = db.query(User).filter(User.tenant_id == tenant_id, User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if payload.full_name is not None:
        user.full_name = payload.full_name
    if payload.birth_date is not None:
        user.birth_date = payload.birth_date
    if payload.tier is not None:
        user.tier = payload.tier
    if payload.bonus_balance is not None:
        user.bonus_balance = int(payload.bonus_balance)
    db.commit()
    db.refresh(user)
    return UserOut.model_validate(user)


# ─── УДАЛЕНИЕ ──────────────────────────────────────────────────────────────────

@router.delete("/{user_id}")
def delete_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    """Удаление клиента. Только owner/admin."""
    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin")
    user = db.query(User).filter(User.tenant_id == tenant_id, User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(user)
    db.commit()
    return {"ok": True, "deleted": user_id}


# ─── GDPR / Анонимизация ───────────────────────────────────────────────────────

@router.post("/gdpr/anonymize")
def anonymize_user_by_phone(
    request: Request,
    phone: str,
    db: Session = Depends(get_db),
):
    """
    GDPR-совместимая анонимизация клиента по номеру телефона.
    Персональные данные заменяются на анонимные, транзакции сохраняются.
    Доступ: owner, admin. Также может использоваться для самоудаления.
    """
    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin")

    phone_clean = normalize_phone(phone)
    user = db.query(User).filter(
        User.tenant_id == tenant_id,
        User.phone == phone_clean,
    ).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    # Анонимизируем данные
    anon_id = f"anon_{user.id}"
    user.phone = anon_id
    user.full_name = "Анонимный клиент"
    user.birth_date = None
    user.bonus_balance = 0
    user.tier = "Bronze"

    db.commit()

    return {
        "ok": True,
        "message": "Client data anonymized. Transactions preserved.",
        "user_id": user.id,
    }


@router.delete("/gdpr/delete")
def gdpr_delete_user(
    request: Request,
    phone: str,
    db: Session = Depends(get_db),
):
    """
    Полное удаление клиента по GDPR-запросу.
    Удаляются все данные: клиент, его транзакции, бонусы.
    Доступ: owner, admin.
    """
    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin")

    phone_clean = normalize_phone(phone)
    user = db.query(User).filter(
        User.tenant_id == tenant_id,
        User.phone == phone_clean,
    ).first()

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    user_id = user.id

    # Удаляем связанные данные
    db.query(Transaction).filter(Transaction.user_id == user_id).delete()
    db.query(BonusGrant).filter(BonusGrant.user_id == user_id).delete()
    db.delete(user)
    db.commit()

    return {
        "ok": True,
        "message": "Client and all associated data deleted (GDPR compliant).",
        "user_id": user_id,
    }


# ─── ЭКСПОРТ ──────────────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_users_excel(request: Request, db: Session = Depends(get_db)):
    """Экспорт всех клиентов тенанта в Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin", "manager")

    rows = (
        db.query(
            User,
            func.coalesce(func.sum(Transaction.paid_amount), 0).label("total_spent"),
            func.count(Transaction.id).label("tx_count"),
        )
        .outerjoin(Transaction, (Transaction.user_id == User.id) & (Transaction.tenant_id == tenant_id))
        .filter(User.tenant_id == tenant_id)
        .group_by(User.id)
        .order_by(User.id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Клиенты"

    headers = ["ID", "Телефон", "Имя", "Дата рождения", "Уровень", "Бонусный баланс", "Сумма покупок", "Кол-во покупок"]
    header_fill = PatternFill("solid", start_color="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    for row_idx, (user, total_spent, tx_count) in enumerate(rows, 2):
        birth = user.birth_date.strftime("%d.%m.%Y") if user.birth_date else ""
        ws.cell(row=row_idx, column=1, value=user.id)
        ws.cell(row=row_idx, column=2, value=user.phone)
        ws.cell(row=row_idx, column=3, value=user.full_name or "")
        ws.cell(row=row_idx, column=4, value=birth)
        ws.cell(row=row_idx, column=5, value=user.tier or "Bronze")
        ws.cell(row=row_idx, column=6, value=user.bonus_balance or 0)
        ws.cell(row=row_idx, column=7, value=int(total_spent))
        ws.cell(row=row_idx, column=8, value=int(tx_count))

        if row_idx % 2 == 0:
            row_fill = PatternFill("solid", start_color="F0F4FA")
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = row_fill

    col_widths = [8, 18, 28, 16, 12, 18, 18, 16]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws2 = wb.create_sheet("Шаблон для импорта")
    tpl_headers = ["Телефон *", "Имя *", "Дата рождения (ДД.ММ.ГГГГ)", "Уровень (Bronze/Silver/Gold)"]
    tpl_fill = PatternFill("solid", start_color="2E7D32")
    tpl_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    for col, h in enumerate(tpl_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = tpl_font
        cell.fill = tpl_fill
        cell.alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 22
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 26
    ws2.column_dimensions["D"].width = 26

    ws2.cell(row=2, column=1, value="77001234567")
    ws2.cell(row=2, column=2, value="Айгуль Иванова")
    ws2.cell(row=2, column=3, value="15.03.1990")
    ws2.cell(row=2, column=4, value="Bronze")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.utcnow().strftime("%Y%m%d")
    filename = f"clients_{today}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─── ИМПОРТ ───────────────────────────────────────────────────────────────────

@router.post("/import/excel")
def import_users_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Импорт клиентов из Excel."""
    import openpyxl

    tenant_id = must_tenant_id(request)
    require_role(request, "owner", "admin", "manager")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Только .xlsx файлы")

    content = file.file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Файл слишком большой (макс 10 МБ)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть файл. Проверьте формат.")

    ws = None
    for sheet_name in wb.sheetnames:
        if "шаблон" not in sheet_name.lower() and "template" not in sheet_name.lower():
            ws = wb[sheet_name]
            break
    if not ws:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Файл пустой")

    header_row = [str(h or "").strip().lower() for h in rows[0]]

    def find_col(*keywords) -> int | None:
        for kw in keywords:
            for i, h in enumerate(header_row):
                if kw in h:
                    return i
        return None

    col_phone = find_col("телефон", "phone", "номер")
    col_name  = find_col("имя", "name", "фио")
    col_birth = find_col("рождени", "birth", "дата")
    col_tier  = find_col("уровень", "tier", "bronze", "статус")

    if col_phone is None:
        raise HTTPException(status_code=400, detail="Колонка 'Телефон' не найдена.")
    if col_name is None:
        raise HTTPException(status_code=400, detail="Колонка 'Имя' не найдена.")

    created = 0
    updated = 0
    skipped = 0
    errors  = []

    VALID_TIERS = {"Bronze", "Silver", "Gold"}

    for row_num, row in enumerate(rows[1:], start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        raw_phone = str(row[col_phone] or "").strip()
        if not raw_phone:
            skipped += 1
            continue
        phone = normalize_phone(raw_phone)
        if len(phone) < 10 or len(phone) > 12:
            errors.append(f"Строка {row_num}: неверный телефон '{raw_phone}'")
            skipped += 1
            continue

        name = str(row[col_name] or "").strip() if col_name is not None else ""
        if not name:
            errors.append(f"Строка {row_num}: имя пустое, пропускаем")
            skipped += 1
            continue

        birth: date | None = None
        if col_birth is not None and row[col_birth]:
            raw_b = str(row[col_birth]).strip()
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
                try:
                    birth = datetime.strptime(raw_b, fmt).date()
                    break
                except ValueError:
                    pass
            if birth is None and hasattr(row[col_birth], "year"):
                try:
                    birth = row[col_birth].date() if hasattr(row[col_birth], "date") else row[col_birth]
                except Exception:
                    pass

        tier = "Bronze"
        if col_tier is not None and row[col_tier]:
            t = str(row[col_tier]).strip().capitalize()
            if t in VALID_TIERS:
                tier = t

        existing = db.query(User).filter(
            User.tenant_id == tenant_id,
            User.phone == phone,
        ).first()

        if existing:
            changed = False
            if name and existing.full_name != name:
                existing.full_name = name
                changed = True
            if birth and existing.birth_date != birth:
                existing.birth_date = birth
                changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            user = User(
                tenant_id=tenant_id,
                phone=phone,
                full_name=name,
                birth_date=birth,
                tier=tier,
                bonus_balance=0,
            )
            db.add(user)
            created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения: {str(e)}")

    return {
        "ok": True,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:20],
        "total_processed": created + updated + skipped,
    }